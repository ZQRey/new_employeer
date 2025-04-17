import openpyxl, os, shutil

def generate_excel_files(data):
    # Путь к шаблонам
    t1 = "templates/templates_doc/pril_eisz.xlsx"
    t2 = "templates/templates_doc/pril_kmis.xlsx"
    wb1, wb2 = openpyxl.load_workbook(t1), openpyxl.load_workbook(t2)
    ws1, ws2 = wb1.active, wb2.active
    # извлечь отделение как прежде
    dn = data['department_dn']
    key = dn.split(",")[0].replace("OU=","").strip().lower()
    disp = key
    ws1["B9"], ws1["C9"], ws1["D9"], ws1["E9"], ws1["G9"] = (
        data['full_name'], data['iin'], data['specialty'], disp, data['phone'])
    ws2["B10"], ws2["C10"], ws2["D10"], ws2["E10"], ws2["F10"], ws2["G10"], ws2["I10"] = (
        data['full_name'], data['iin'], data['email'], data['phone'], data['specialty'], disp, data['specialty'])
    suf = f"{data['last_name']}_{data['first_name']}".replace(" ","_")
    out1 = f"static/Doc/EISZ_DOC/pril_eisz_{suf}.xlsx"
    out2 = f"static/Doc/KMIS_DOC/pril_kmis_{suf}.xlsx"
    os.makedirs(os.path.dirname(out1), exist_ok=True)
    os.makedirs(os.path.dirname(out2), exist_ok=True)
    wb1.save(out1); wb2.save(out2)
    return out1.split("static/")[1], out2.split("static/")[1]

def generate_obaz_excel(data):
    tpl = "templates/templates_doc/obaz_eisz.xlsx"
    wb = openpyxl.load_workbook(tpl)
    ws = wb.active
    ws["B10"] = data['full_name']
    ws["A13"] = data['specialty']
    suf = f"{data['last_name']}_{data['first_name']}".replace(" ","_")
    out = f"static/Doc/OBAZ_DOC/obaz_eisz_{suf}.xlsx"
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out.split("static/")[1]

def copy_obaz_docx(data):
    src = "templates/templates_doc/obaz_eisz.docx"
    suf = f"{data['last_name']}_{data['first_name']}".replace(" ","_")
    out = f"static/Doc/OBAZ_DOC/obaz_eisz_{suf}.docx"
    os.makedirs(os.path.dirname(out), exist_ok=True)
    shutil.copy(src, out)
    return out.split("static/")[1]
